import requests
import openpyxl
import csv
from bs4 import BeautifulSoup as bs


def write_xls(lst):
    wb = openpyxl.Workbook()
    wb.create_sheet(title="Лист", index=0)
    sheet = wb["Лист"]
    for i in range(len(lst)):
        for j in lst[i]:
            value = str(j)
            cell = sheet.cell(row=i + 1, column=(lst[i].index(j)) + 1)
            cell.value = value
    wb.save('C:\\Users\\YumagulovA\\Desktop\\ozon\\motul_ozon_add.xlsx')


def excel_reader(file_obj):
    result = []
    wb = openpyxl.load_workbook(file_obj)
    sheet = wb['МП']
    for row in sheet['A']:
        result.append(str(row.value))
        print(row.value)
    return result


def csv_reader(file_obj):
    """
    Read a csv file
    """
    reader = csv.reader(file_obj)
    for row in reader:
        if motul_image.isdigit() and len(motul_image) <= 6:
            motul_image = row[0].split(';')[0]
            print(motul_image)


def get_link(url):
    r = requests.get(url)
    if r.status_code == 200:
        soup = bs(r.text, "html.parser")
        divs = soup.find_all('div', attrs={"class": "catalog_content_products-item"})
        for div in divs:
            product = div.find('span', attrs={"class": 'catalog-card-capacity__liter active-liter'})
            image = product.get('data-image')
            image_list = (image.split('/'))
            # print(image_list)
            correct_url = f"https://motul.store/{image_list[1]}/{image_list[2]}/{image_list[3]}/{image_list[4]}/" \
                          f"700_700_1/{image_list[-1]}"
            return correct_url

    else:
        print(r.status_code)


csv_path = "C:\\Users\\YumagulovA\\Desktop\\ozon\\products.csv"
xls_path = "C:\\Users\\YumagulovA\\Desktop\\ozon\\motul.xlsx"
lst = []
with open(xls_path, "r") as f_obj:
    reader = excel_reader(f_obj)
    for row in reader:
        motul_image = row[0].split(';')[0]
        if motul_image.isdigit() and len(motul_image) <= 6:
            with open("C:\\Users\\YumagulovA\\Desktop\\ozon\\file.txt", 'a') as f:
                correct_link = get_link(f"https://motul.store/search/?q={motul_image}")
                if correct_link is None:
                    continue
                else:
                    print(correct_link)
                    f.write(correct_link + '\n')
