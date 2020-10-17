import openpyxl
import requests
from bs4 import BeautifulSoup as bs


# wb = openpyxl.load_workbook("C:\\Users\\YumagulovA\\Desktop\\ozon\\motul.xlsx")
# # print(wb.sheetnames)
# sheet = wb['МП']
# # print(sheet[f'A{}'].value)
# i = 0
# for row in sheet['A']:
#     print(row.value)
#     i += 1
#     print(i)

def excel_reader(file_obj):
    result = []
    wb = openpyxl.load_workbook(file_obj)
    sheet = wb['МП']
    for row in sheet['A']:
        result.append(str(row.value))
        print(row.value)
    return result


def get_link(url):
    r = requests.get(url)
    if r.status_code == 200:
        soup = bs(r.text, "html.parser")
        divs = soup.find_all('div', attrs={"class": "catalog_content_products-item"})
        for div in divs:
            # try:
                # product = div.find('span', attrs={"class": 'catalog-card-capacity__liter active-liter'})
                product = div.find('img', attrs={"class": 'catalog_content_products-img'})
                image = product.get('data-image')
            # except AttributeError:
                product = div.find('img', attrs={"class": 'catalog_content_products-img'})
                print(product)
                image = product.get("src")
                if len(image) != 0:
                    image_list = (image.split('/'))
                    correct_url = f"https://motul.store/{image_list[1]}/{image_list[2]}/{image_list[3]}/{image_list[4]}/" \
                                  f"700_700_1/{image_list[-1]}"
                    print("exept", correct_url)
                    return correct_url
                else:
                    return None
            # else:
                # product = div.find('img', attrs={"class": 'catalog_content_products-img'})
                print(product)
                image = product.get("src")
                if len(image) != 0:
                    image_list = (image.split('/'))
                    correct_url = f"https://motul.store/{image_list[1]}/{image_list[2]}/{image_list[3]}/{image_list[4]}/" \
                                  f"700_700_1/{image_list[-1]}"
                    print("else", correct_url)
                    return correct_url
                else:
                    return None


    else:
        print(r.status_code)


count = 1
xls_path = "C:\\Users\\YumagulovA\\Desktop\\ozon\\motul.xlsx"
with open("C:\\Users\\YumagulovA\\Desktop\\ozon\\file.txt", 'a') as f:
    for i in excel_reader(xls_path):
        correct_link = get_link(f"https://motul.store/search/?q={i}")
        if correct_link != None:
            print(count)
            # print(correct_link)
            count += 1
            # f.write(correct_link + '\n')

