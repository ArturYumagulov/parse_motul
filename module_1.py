import openpyxl
import requests
from bs4 import BeautifulSoup as bs


def write_xls(lst, path):
    wb = openpyxl.Workbook()
    wb.create_sheet(title="Лист", index=0)
    sheet = wb["Лист"]
    for i in range(len(lst)):
        for j in lst[i]:
            value = str(j)
            cell = sheet.cell(row=i + 1, column=(lst[i].index(j)) + 1)
            cell.value = value
    wb.save(path)


def excel_reader(file_obj):
    result = []
    wb = openpyxl.load_workbook(file_obj)
    sheet = wb['МП']
    for row in sheet['A']:
        result.append(str(row.value))
    return result


def get_link(url):
    r = requests.get(url)
    price = None
    vendor_code = None
    if r.status_code == 200:
        soup = bs(r.text, "html.parser")
        divs = soup.find_all('div', attrs={"class": "catalog_content_products-item"})
        for div in divs:
            # print(div)
            types = div.find('span', attrs={'class': 'catalog_content_products-category'}).text
            if types in ["Моторные масла Motul", "Трансмиссионные масла Motul", "Моторные масла 300V Motul"]:
                product = div.find('img', attrs={"class": 'catalog_content_products-img'})
                litres = div.find('span', attrs={'class': 'catalog-card-capacity__liter active-liter'})
                name = div.find('div', attrs={'class': 'catalog_content_products-text'}).text.replace("   ", '').\
                                                                                            replace('\n', '')
                if "Трансмиссионные масла" in types:
                    types = "Автохимия. Трансмиссионное, гидравлическое масла"
                    name = "Трансмиссионное масло Motul" + name[28:-6]
                elif "Моторные масла" in types:
                    types = "Автохимия. Масло моторное"
                    name = "Моторное масла Motul" + name[21:-6]
                if litres is not None:
                    price = litres.get('data-price')
                    vendor_code = litres.get('data-articul')
                image = product.get("src")
                weight = 0
                length = 0
                height = 0
                if "5" in litres.text.strip():
                    litres = 5000
                    weight = 150
                    length = 200
                    height = 300
                elif "4" in litres.text.strip():
                    litres = 4000
                    weight = 150
                    length = 200
                    height = 300
                elif "2" in litres.text.strip():
                    litres = 2000
                    weight = 10
                    length = 10
                    height = 20
                elif "1" in litres.text.strip():
                    litres = 1000
                    weight = 10
                    length = 10
                    height = 20
                if len(image) != 0:
                    image_list = (image.split('/'))
                    correct_url = f"https://motul.store/{image_list[1]}/{image_list[2]}/{image_list[3]}/{image_list[4]}/" \
                                  f"700_700_1/{image_list[-1]}"
                    return [vendor_code, name, int(price.replace(" руб.", "").replace(' ', '')), types, litres, weight,
                            length, height, correct_url]
                else:
                    return None
            else:
                continue
    else:
        print(r.status_code)


if __name__ == '__main__':
    get_link(f"https://motul.store/search/?q={108945}")
    #  .catalog_content_products-text > br:nth-child(2)
