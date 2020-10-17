import openpyxl

lst = [["test1", 'A'],
       ["test2", 'B'],
       ["test3", "C"],
       ["test4", 'D'],
       ["test5", "R"]]

def write_xls(lst):
    wb = openpyxl.Workbook()
    wb.create_sheet(title="Лист", index=0)
    sheet = wb["Лист"]
    for i in range(len(lst)):
        for j in lst[i]:
            value = str(j)
            cell = sheet.cell(row=i + 1, column=(lst[i].index(j)) + 1)
            cell.value = value
    wb.save('/home/zico/Desktop/example.xlsx')


# write_xls(lst)


